from image_scrap import *
from openpyxl import load_workbook
import os, time, shutil
from walk import walk_folder
print("hello")
loc  = 'status female_latest.xlsx'
wb = load_workbook(loc)
# name_list = ["ambati rayudu","Vikas gowda"]
# folder_num = [1,2,3]
name_list = []
folder_num = []
sheet = wb['female1']
for i in range(930,955):
    if sheet.cell(row=i, column=4).value:
        # print(str(sheet.cell(row=i, column=2).value)+","+sheet.cell(row=i, column=4).value+", "+sheet.cell(row=i, column=7).value)
        folder_num.append(sheet.cell(row=i, column=2).value)
        name_list.append(sheet.cell(row=i, column=4).value+", "+sheet.cell(row=i, column=7).value)
    else:
        break


i=1
def check(file_c, num_img, pname):
    number_of_images = num_img
    file_count = file_c
    person = pname
    if file_count<40 and number_of_images<170:
        print("Re Downloading.........Hi Anurag I AM DOWNLOADING.............",person,number_of_images)
        time.sleep(2)
        
        shutil.rmtree(SRC_PATH)
        shutil.rmtree(DST_PATH)
        crawl(SRC_PATH, name, number_of_images)
        walk_folder(SRC_PATH,DST_PATH)
        path, dirs, files = next(os.walk(DST_PATH))
        file_count = len(files)
        print(file_count,DST_PATH)
        number_of_images +=40
        time.sleep(2)
        check(file_count, number_of_images, person)

for name,num in zip(name_list,folder_num):
    SRC_PATH = './female_download_thursday_25/'+str(num).zfill(4)+'/'
    DST_PATH = './female_crop_status_thursday_25/'+str(num).zfill(4)+'/'
    number_of_images = 80
    print("Downloading........Yeaaaaaa..............",name)
    crawl(SRC_PATH, name, number_of_images)
    walk_folder(SRC_PATH,DST_PATH)
    path, dirs, files = next(os.walk(DST_PATH))
    file_count = len(files)
    print(file_count,DST_PATH)
    number_of_images +=40
    time.sleep(2)
    check(file_count, number_of_images, name)
    i+=1
    # time.sleep(2)
